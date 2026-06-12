"""Import khách CRM đa nguồn — parse file + auto-detect cột + chuẩn hoá dòng.

Dùng chung cho 2 nguồn (Google Sheet & file upload). Luồng:
  1. Nguồn trả về bảng thô: list[list[str]] (dòng đầu = header).
  2. table_to_records() → headers + list[dict] (mỗi dict key theo header).
  3. suggest_mapping(headers) → gợi ý cột nào là tên/sđt/email/nguồn/ghi chú/nhu cầu.
  4. FE cho admin chỉnh mapping → commit.
  5. records_to_leads(records, mapping) → list[{name, phone, email, note}] sạch.
  6. lead_store.import_customers(...) dedupe + tạo + (tuỳ chọn) auto-assign + auto-care.

Parse file: CSV (module csv chuẩn) + XLSX (openpyxl, đã có trong venv). KHÔNG phụ
thuộc pandas. Bỏ dấu tiếng Việt khi so khớp tên cột để auto-detect chính xác.
"""

from __future__ import annotations

import csv
import io
import logging
import unicodedata
from typing import Optional

log = logging.getLogger(__name__)

# Khoá nội bộ gắn TÊN TAB vào từng dòng (không phải header thật, không hiện trong
# UI map cột). records_to_leads đọc khoá này để gắn nhãn vùng miền / tệp khách.
TAB_KEY = "__tab__"

# Trường chuẩn hệ thống → từ khoá header (đã bỏ dấu, lowercase) để auto-detect.
#
# THỨ TỰ QUAN TRỌNG: mỗi header chỉ gán cho 1 field (field đứng trước thắng). Đặt
# các field phân loại CỤ THỂ (region/customer_group/project/product_type/budget/
# purpose) TRƯỚC `demand`/`note` chung để cột "Sản phẩm quan tâm", "Ngân sách"...
# không bị `demand` chiếm mất.
_FIELD_KEYWORDS: dict[str, tuple[str, ...]] = {
    "name": ("ten", "name", "hovaten", "ho ten", "ho va ten", "khachhang",
             "khach hang", "fullname", "full name", "customer"),
    "phone": ("sdt", "sodienthoai", "so dien thoai", "phone", "dienthoai",
              "dien thoai", "mobile", "tel", "lien he", "lienhe", "phone number"),
    "email": ("email", "mail", "e-mail", "thu dien tu"),
    # Vùng miền / khu vực.
    "region": ("vung mien", "vungmien", "vung", "mien", "khu vuc", "khuvuc",
               "region", "area", "tinh", "tinh thanh", "thanh pho", "province",
               "city", "dia ban", "diaban"),
    # Tệp khách / nhóm khách (cũng là nhãn tab khi import nhiều tab).
    "customer_group": ("tep khach", "tepkhach", "nhom khach", "nhomkhach", "nhom",
                       "group", "customer group", "tep", "tep data", "phan loai khach",
                       "nhom data", "segment", "phan khuc khach"),
    # Dự án quan tâm.
    "project": ("du an", "duan", "project", "du an quan tam"),
    # Phân khúc / sản phẩm quan tâm (liền kề / shophouse / căn hộ...).
    "product_type": ("san pham quan tam", "san pham", "sanpham", "loai san pham",
                     "loai sp", "phan khuc", "phankhuc", "product", "product type",
                     "loai hinh", "loai bds", "loai hinh san pham", "lien ke",
                     "shophouse", "can ho", "biet thu", "dat nen"),
    # Ngân sách / khả năng tài chính.
    "budget": ("ngan sach", "ngansach", "budget", "tai chinh", "kha nang tai chinh",
               "muc gia", "tam gia", "khoang gia", "price range", "tam tien"),
    # Mục đích (ở / đầu tư).
    "purpose": ("muc dich", "mucdich", "purpose", "nhu cau o", "de o", "mua de o",
                "dau tu", "muc dich mua", "muc dich su dung"),
    "source": ("nguon", "source", "kenh", "channel", "origin"),
    "note": ("ghichu", "ghi chu", "note", "notes", "remark", "remarks", "comment"),
    "demand": ("nhucau", "nhu cau", "demand", "need", "needs", "yeu cau", "yeucau",
               "quan tam", "quantam", "interest"),
}

CANONICAL_FIELDS = list(_FIELD_KEYWORDS.keys())


def _strip_accents(text: str) -> str:
    nfkd = unicodedata.normalize("NFKD", text or "")
    no_accent = "".join(c for c in nfkd if not unicodedata.combining(c))
    return no_accent.replace("đ", "d").replace("Đ", "D")


def _norm_header(text: str) -> str:
    return _strip_accents((text or "").strip().lower())


# ---------------------------------------------------------------------------
# Parse file bytes → bảng thô
# ---------------------------------------------------------------------------

def parse_csv(content: bytes) -> list[list[str]]:
    """Parse CSV bytes → list[list[str]]. Tự đoán separator (, ; \\t)."""
    text = _decode(content)
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel  # mặc định dấu phẩy
    reader = csv.reader(io.StringIO(text), dialect)
    rows = [[(c or "").strip() for c in row] for row in reader if any(row)]
    return rows


def parse_xlsx(content: bytes) -> list[list[str]]:
    """Parse XLSX bytes (sheet đầu) → list[list[str]] qua openpyxl."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        cells = ["" if c is None else str(c).strip() for c in row]
        if any(cells):
            rows.append(cells)
    wb.close()
    return rows


def parse_xlsx_sheets(content: bytes) -> list[tuple[str, list[list[str]]]]:
    """Parse XLSX bytes → [(tên_sheet, list[list[str]]), ...] cho MỌI sheet có dữ
    liệu. Dùng cho import nhiều tab từ file Excel. Sheet rỗng được bỏ qua."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    out: list[tuple[str, list[list[str]]]] = []
    for ws in wb.worksheets:
        rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            cells = ["" if c is None else str(c).strip() for c in row]
            if any(cells):
                rows.append(cells)
        if rows:
            out.append((ws.title, rows))
    wb.close()
    return out


def is_multi_sheet_xlsx(filename: str, content: bytes) -> bool:
    """True nếu file là .xlsx/.xlsm và có >1 sheet chứa dữ liệu."""
    name = (filename or "").lower()
    if not (name.endswith(".xlsx") or name.endswith(".xlsm")):
        return False
    try:
        return len(parse_xlsx_sheets(content)) > 1
    except Exception:  # noqa: BLE001 — không đọc được → coi như đơn sheet
        return False


def parse_file(filename: str, content: bytes) -> list[list[str]]:
    """Chọn parser theo đuôi file. Raise ValueError nếu định dạng không hỗ trợ."""
    name = (filename or "").lower()
    if name.endswith(".csv") or name.endswith(".tsv") or name.endswith(".txt"):
        return parse_csv(content)
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return parse_xlsx(content)
    raise ValueError(
        f"Định dạng file không hỗ trợ: {filename}. Chỉ nhận .csv hoặc .xlsx."
    )


def _decode(content: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "cp1258", "latin-1"):
        try:
            return content.decode(enc)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


# ---------------------------------------------------------------------------
# Bảng thô → records + auto-detect mapping
# ---------------------------------------------------------------------------

def table_to_records(rows: list[list[str]]) -> tuple[list[str], list[dict]]:
    """Dòng đầu = header. Trả (headers, list[dict] theo header). Header rỗng được
    đặt tên cot_1, cot_2... để không mất cột."""
    if not rows:
        return [], []
    raw_header = rows[0]
    headers: list[str] = []
    seen: dict[str, int] = {}
    for i, h in enumerate(raw_header):
        name = (h or "").strip() or f"cot_{i + 1}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        headers.append(name)
    records: list[dict] = []
    for row in rows[1:]:
        rec = {headers[i]: (row[i] if i < len(row) else "") for i in range(len(headers))}
        if any((v or "").strip() for v in rec.values()):
            records.append(rec)
    return headers, records


def merge_tabbed_records(
    tables: list[tuple[str, list[list[str]]]]
) -> tuple[list[str], list[dict], list[dict]]:
    """Gộp nhiều tab (tên_tab, bảng_thô) → (headers_hợp_nhất, records, tab_counts).

    - headers: hợp (union) header của mọi tab, giữ thứ tự xuất hiện đầu tiên.
    - records: mọi dòng của mọi tab; mỗi dòng được gắn TAB_KEY = tên tab.
    - tab_counts: [{"name": tên_tab, "count": số_dòng}] để FE báo "số dòng/tab".
    """
    headers: list[str] = []
    seen: set[str] = set()
    records: list[dict] = []
    tab_counts: list[dict] = []
    for tab_name, table in tables:
        h_i, recs_i = table_to_records(table)
        for h in h_i:
            if h not in seen:
                seen.add(h)
                headers.append(h)
        for rec in recs_i:
            rec[TAB_KEY] = tab_name
        records.extend(recs_i)
        tab_counts.append({"name": tab_name, "count": len(recs_i)})
    return headers, records, tab_counts


def suggest_mapping(headers: list[str]) -> dict[str, Optional[str]]:
    """Gợi ý mapping {field_chuẩn: header}. Header nào khớp nhiều field → field
    đầu thắng; mỗi header chỉ gán cho 1 field."""
    mapping: dict[str, Optional[str]] = {f: None for f in CANONICAL_FIELDS}
    used: set[str] = set()
    norm = {h: _norm_header(h) for h in headers}
    for field, keywords in _FIELD_KEYWORDS.items():
        for h in headers:
            if h in used:
                continue
            nh = norm[h]
            if any(kw == nh for kw in keywords) or any(kw in nh for kw in keywords):
                mapping[field] = h
                used.add(h)
                break
    return mapping


# ---------------------------------------------------------------------------
# Records + mapping → lead dicts sạch
# ---------------------------------------------------------------------------

# Trường phân loại MỞ RỘNG được map thẳng (1 cột → 1 field, không gộp vào note).
_PROFILE_MAP_FIELDS = (
    "region", "customer_group", "product_type", "budget", "purpose", "project",
)


def records_to_leads(
    records: list[dict], mapping: dict[str, Optional[str]]
) -> list[dict]:
    """Áp mapping → list[dict] lead sạch (name/phone/email/note + trường phân loại).

    - 'demand' (nhu cầu tự do) vẫn gộp vào note dạng 'Nhu cầu: ...' để AI scoring/
      insight (Phần B) dùng được ngay (giữ tương thích ngược).
    - Các trường phân loại (region/customer_group/product_type/budget/purpose/
      project) map THẲNG sang field riêng — chỉ thêm khi có giá trị.
    - Nhãn TAB (TAB_KEY): nếu dòng thuộc 1 tab và CHƯA có region/customer_group
      riêng → lấy tên tab làm vùng miền + tệp khách (ưu tiên giá trị riêng của dòng).
    """
    def col(field: str) -> Optional[str]:
        return mapping.get(field)

    def val(rec: dict, field: str) -> str:
        c = col(field)
        return (rec.get(c) or "").strip() if c else ""

    name_col = col("name")
    phone_col = col("phone")
    email_col = col("email")
    note_col = col("note")
    demand_col = col("demand")
    src_col = col("source")

    leads: list[dict] = []
    for rec in records:
        note_parts: list[str] = []
        if note_col and (rec.get(note_col) or "").strip():
            note_parts.append(rec[note_col].strip())
        if demand_col and (rec.get(demand_col) or "").strip():
            note_parts.append(f"Nhu cầu: {rec[demand_col].strip()}")
        lead: dict = {
            "name": (rec.get(name_col) or "").strip() if name_col else "",
            "phone": (rec.get(phone_col) or "").strip() if phone_col else "",
            "email": (rec.get(email_col) or "").strip() if email_col else "",
            "note": " — ".join(note_parts) or None,
        }
        # Trường phân loại map thẳng (chỉ khi có giá trị).
        for field in _PROFILE_MAP_FIELDS:
            v = val(rec, field)
            if v:
                lead[field] = v
        # Nhãn tab → vùng miền / tệp khách (giá trị riêng của dòng được ưu tiên).
        tab = (rec.get(TAB_KEY) or "").strip()
        if tab:
            lead.setdefault("customer_group", tab)
            lead.setdefault("region", tab)
        # Nguồn theo từng dòng (nếu cột source được map) — ghi đè nguồn mặc định.
        if src_col and (rec.get(src_col) or "").strip():
            lead["row_source"] = rec[src_col].strip()
        leads.append(lead)
    return leads
